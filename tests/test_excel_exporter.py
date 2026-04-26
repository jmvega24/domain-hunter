from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from domainhunter.exporters.excel_exporter import export_results_to_excel
from domainhunter.models import Confidence, DomainCheckResult, DomainStatus


def test_export_results_to_excel_writes_expected_columns(tmp_path: Path) -> None:
    output = tmp_path / "results.xlsx"
    result = DomainCheckResult(
        domain="clavora.com",
        provider="godaddy",
        status=DomainStatus.MANUAL_REVIEW,
        checked_at=datetime(2026, 4, 26, tzinfo=timezone.utc),
        confidence=Confidence.LOW,
        notes="Sin evidencia suficiente.",
    )

    export_results_to_excel([result], output)

    workbook = load_workbook(output)
    sheet = workbook.active

    assert workbook.sheetnames == ["results", "summary"]
    assert [cell.value for cell in sheet[1]] == [
        "domain",
        "provider",
        "status",
        "price",
        "currency",
        "checked_at",
        "confidence",
        "notes",
        "error_message",
    ]
    assert sheet["A2"].value == "clavora.com"
    assert sheet["C2"].value == "manual_review"

    summary = workbook["summary"]
    assert summary["A2"].value == "clavora.com"
    assert summary["B2"].value == "manual_review"
